import openpyxl

def getDataofHeader(xlFile, sheet, header, headerRow = 1):
    wb = openpyxl.load_workbook(xlFile)
    mainSheet = wb[sheet]
    
    dataUnderTheHeader = []
    maxcolumn = mainSheet.max_column
    mincolumn = mainSheet.min_column
    maxrow = mainSheet.max_row
    minrow = mainSheet.min_row
    for i in range(mincolumn, maxcolumn):
        if mainSheet.cell(row=headerRow,column=i).value == header:
            for j in range(minrow+1, maxrow+1):
                dataUnderTheHeader.append(mainSheet.cell(row=j,column=i).value)
            break
    return dataUnderTheHeader
    

if __name__ == '__main__':
    print(getDataofHeader('fruits.xlsx', 'Main', 'Quantity'))
